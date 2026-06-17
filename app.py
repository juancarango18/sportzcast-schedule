import streamlit as st
import pandas as pd
import json
import os
import psycopg2
import bcrypt
import calendar
import openpyxl
import scheduler
import sys
import io
from datetime import datetime, timedelta, date

# ==========================================
# ENTERPRISE UI SETUP (MUST BE FIRST!)
# ==========================================
st.set_page_config(
    page_title="Sportzcast Scheduler",
    page_icon="🗓️", 
    layout="wide", # This stretches the app to use the whole screen!
    initial_sidebar_state="expanded"
)

try:
    st.logo("logo.png")
except:
    pass # Fails safely if the logo hasn't loaded yet

# ==========================================
# PAGE CONFIGURATION
# ==========================================
st.set_page_config(page_title="Sports Scheduler Pro", page_icon="🗓️", layout="wide")

TEAM_MEMBERS = ["Alberto Salazar", "Camilo Buritica", "Emilio Gonzalez", "Juan Camilo Correa", "Simon Mejia", "Brayan Carlosama"]
GAMES_SCHEDULE_PATH = "games_schedule.csv"
REQUIRED_GAMES_COLUMNS = ["Date", "Sport", "Matchup", "Coverage_Start", "Coverage_End", "Venue"]

def save_games_schedule_upload(uploaded_file):
    uploaded_file.seek(0)
    df = pd.read_csv(uploaded_file)
    missing = [col for col in REQUIRED_GAMES_COLUMNS if col not in df.columns]
    if missing:
        return False, f"Missing required columns: {', '.join(missing)}"
    if df.empty:
        return False, "CSV file is empty."
    df.to_csv(GAMES_SCHEDULE_PATH, index=False)
    return True, f"Saved {len(df)} games to {GAMES_SCHEDULE_PATH}"

# ==========================================
# DATABASE HELPER FUNCTIONS (POSTGRES CLOUD)
# ==========================================
def get_db_connection():
    return psycopg2.connect(st.secrets["connections"]["supabase"]["url"])

def verify_login(identifier, password):
    conn = get_db_connection()
    c = conn.cursor()
    # Check if they typed their username OR their email
    c.execute("SELECT password_hash, role, username FROM users WHERE username=%s OR email=%s", (identifier, identifier))
    result = c.fetchone()
    conn.close()
    if result and bcrypt.checkpw(password.encode('utf-8'), result[0].encode('utf-8')):
        return result[1], result[2] # Returns (role, actual_username)
    return None, None

def save_user_requests(username, year, month, ptos, rdos):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("DELETE FROM time_off_requests WHERE username=%s AND target_year=%s AND target_month=%s", (username, year, month))
    
    for p in ptos:
        c.execute("INSERT INTO time_off_requests (username, request_date, request_type, target_year, target_month) VALUES (%s, %s, 'PTO', %s, %s)", (username, p, year, month))
    for r in rdos:
        c.execute("INSERT INTO time_off_requests (username, request_date, request_type, target_year, target_month) VALUES (%s, %s, 'RDO', %s, %s)", (username, r, year, month))
        
    conn.commit()
    conn.close()

def get_all_requests(year, month, username=None):
    conn = get_db_connection()
    c = conn.cursor()
    
    if username:
        c.execute("SELECT request_date, request_type FROM time_off_requests WHERE target_year=%s AND target_month=%s AND username=%s", (year, month, username))
        rows = c.fetchall()
        conn.close()
        pto = [r[0] for r in rows if r[1] == 'PTO']
        rdo = [r[0] for r in rows if r[1] == 'RDO']
        return pto, rdo
    else:
        c.execute("SELECT username, request_date, request_type FROM time_off_requests WHERE target_year=%s AND target_month=%s", (year, month))
        rows = c.fetchall()
        conn.close()
        pto_dict, rdo_dict = {}, {}
        for r in rows:
            if r[2] == 'PTO':
                pto_dict.setdefault(r[0], []).append(r[1])
            else:
                rdo_dict.setdefault(r[0], []).append(r[1])
        return pto_dict, rdo_dict

# ==========================================
# SESSION STATE (LOGIN TRACKING)
# ==========================================
if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.username = ''
    st.session_state.role = ''

def logout():
    st.session_state.logged_in = False
    st.session_state.username = ''
    st.session_state.role = ''
    #st.rerun()

# ==========================================
# UI: LOGIN & SIGN-UP SCREEN
# ==========================================
if not st.session_state.logged_in:
    # We use columns to center the login box perfectly in the middle of the screen
    spacer_left, center_col, spacer_right = st.columns([1, 2, 1])
    
    with center_col:
        try:
            st.image("logo.png", width=250) # You can adjust this width!
        except:
            pass
            
        st.title("Support Scheduler")
        st.markdown("Welcome! Please log in or claim your team account.")
        
        tab1, tab2 = st.tabs(["🔒 Log In", "📝 Sign Up / Claim Account"])
        
        with tab1:
            with st.form("login_form"):
                user_input = st.text_input("Email or Username")
                pass_input = st.text_input("Password", type="password")
                submit_login = st.form_submit_button("Log In", use_container_width=True)
                
                if submit_login:
                    role, actual_username = verify_login(user_input, pass_input)
                    if role:
                        st.session_state.logged_in = True
                        st.session_state.username = actual_username
                        st.session_state.role = role
                        st.rerun()
                    else:
                        st.error("Incorrect email/username or password.")
                        
        with tab2:
            st.info("First time here? Select your name to set up your account.")
            with st.form("signup_form"):
                new_user = st.selectbox("Select Your Name", TEAM_MEMBERS)
                new_email = st.text_input("Email Address")
                new_pass = st.text_input("Create a Password", type="password")
                confirm_pass = st.text_input("Confirm Password", type="password")
                submit_signup = st.form_submit_button("Create Account", use_container_width=True)
                
                # ... keep your existing sign-up logic inside here ...
                if submit_signup:
                    if not new_email:
                        st.error("Please enter an email address.")
                    elif new_pass != confirm_pass:
                        st.error("Passwords do not match!")
                    elif len(new_pass) < 6:
                        st.error("Password must be at least 6 characters.")
                    else:
                        try:
                            hashed_pw = bcrypt.hashpw(new_pass.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                            conn = get_db_connection()
                            c = conn.cursor()
                            c.execute("UPDATE users SET password_hash=%s, email=%s WHERE username=%s", (hashed_pw, new_email, new_user))
                            conn.commit()
                            conn.close()
                            st.success("✅ Account successfully claimed! You can now log in using your Email.")
                        except psycopg2.errors.UniqueViolation:
                            st.error("That email is already registered to another account.")
                        
    st.stop()

# ==========================================
# UI: TOP NAVIGATION & CLOUD FETCHING
# ==========================================
st.sidebar.markdown(f"👤 Logged in as: **{st.session_state.username}**")
st.sidebar.button("Logout", on_click=logout)
st.sidebar.markdown("---")

st.title("🗓️ Sportzcast Scheduling Matrix")

col_y, col_m = st.columns(2)
with col_y:
    selected_year = st.selectbox("Select Year", [2026, 2027, 2028])
with col_m:
    selected_month = st.selectbox("Select Month", list(range(1, 13)), index=4, format_func=lambda x: calendar.month_name[x])

days_in_month = calendar.monthrange(selected_year, selected_month)[1]
available_dates = [f"{selected_year}-{str(selected_month).zfill(2)}-{str(i).zfill(2)}" for i in range(1, days_in_month + 1)]

today = datetime.today()
deadline_passed = False
if selected_year < today.year or (selected_year == today.year and selected_month < today.month):
    deadline_passed = True
elif selected_year == today.year and selected_month == today.month + 1 and today.day > 25:
    deadline_passed = True

# FETCH EXCEL FILE DIRECTLY FROM SUPABASE!
conn = get_db_connection()
c = conn.cursor()
c.execute("SELECT is_approved, excel_file FROM schedule_status WHERE target_year=%s AND target_month=%s", (selected_year, selected_month))
status_row = c.fetchone()
conn.close()

is_approved = False
db_file_bytes = None
if status_row:
    is_approved = status_row[0]
    if status_row[1]:
        db_file_bytes = bytes(status_row[1])

# ==========================================
# UI: TEAM MEMBER PORTAL
# ==========================================
elif st.session_state.role == "user":
        try:
            st.image("logo2.png", width=200)
        except:
            pass
            
        st.title(f"👋 Welcome, {st.session_state.username}!")

if st.session_state.role == 'user':
    
    if db_file_bytes and is_approved:
        st.success(f"🎉 The schedule for {calendar.month_name[selected_month]} {selected_year} has been approved and published!")
        st.markdown(f"### 📊 Final Schedule")
        try:
            preview_df = pd.read_excel(io.BytesIO(db_file_bytes), index_col=0).fillna("") 
            st.dataframe(preview_df, use_container_width=True)
        except Exception as e:
            st.error("Could not load preview.")
            
        st.download_button(
            label="📥 Download Approved Matrix",
            data=db_file_bytes,
            file_name=f"APPROVED_{calendar.month_name[selected_month]}_{selected_year}_Matrix.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        st.markdown("---")
        
    st.header("📝 Submit Your Time Off")
    
    if is_approved:
        st.info("The schedule has been finalized. Requests are now locked.")
        disabled = True
    elif deadline_passed:
        st.error(f"The deadline to submit requests for {calendar.month_name[selected_month]} {selected_year} has passed. Please contact your Team Lead.")
        disabled = True
    else:
        st.info(f"Deadline to submit: {calendar.month_name[selected_month - 1 if selected_month > 1 else 12]} 25th")
        disabled = False
        
    current_ptos, current_rdos = get_all_requests(selected_year, selected_month, st.session_state.username)
    
    with st.form("request_form"):
        user_pto = st.multiselect("🏖️ Paid Time Off (PTO)", options=available_dates, default=current_ptos, disabled=disabled)
        user_rdo = st.multiselect("🔄 Requested Day Off (RDO)", options=available_dates, default=current_rdos, disabled=disabled)
        submitted = st.form_submit_button("Save Requests", disabled=disabled)
        
        if submitted:
            save_user_requests(st.session_state.username, selected_year, selected_month, user_pto, user_rdo)
            st.success("Your requests have been successfully saved to the database!")

# ==========================================
# UI: TEAM LEAD (ADMIN) PORTAL
# ==========================================
# ==========================================
#if st.session_state.role == "admin":
        
        # 1. The Logo goes FIRST!
        #st.image("logo2.png", width=200)
        
        # 2. The Main Title goes SECOND! (Delete the duplicate one if you have two!)
        #st.title("🗓️ Sportzcast Scheduler")
        
        # 3. Your Selectors go THIRD!
        #col_year, col_month = st.columns(2)
        #with col_year:
            #selected_year = st.selectbox("Select Year", [2024, 2025, 2026])
        
if st.session_state.role == 'admin':
    
    all_ptos, all_rdos = get_all_requests(selected_year, selected_month)
    
    with st.sidebar:
        st.header("⚙️ Admin Controls")
        if is_approved:
            st.error("🔒 SCHEDULE LOCKED")
            st.markdown("This schedule has been approved. Unlock to regenerate.")
            if st.button("🔓 Unlock Schedule"):
                conn = get_db_connection()
                c = conn.cursor()
                c.execute("UPDATE schedule_status SET is_approved=FALSE WHERE target_year=%s AND target_month=%s", (selected_year, selected_month))
                conn.commit()
                conn.close()
                st.rerun()
        else:
            st.success("🟢 SCHEDULE OPEN")
            st.markdown("This schedule is currently in draft mode.")
            
        st.markdown("---")
        st.subheader("🇨🇴 Colombian Holidays")
        holiday_dates = st.multiselect("Select Holiday Dates", options=available_dates, disabled=is_approved)
        holiday_workers = {}
        if holiday_dates:
            for h_date in holiday_dates:
                workers = st.multiselect(f"Who is WORKING on {h_date}?", options=TEAM_MEMBERS, max_selections=3, disabled=is_approved)
                holiday_workers[h_date] = workers
                
        with st.expander("👀 View Team Requests"):
            st.json({"PTO": all_ptos, "RDO": all_rdos})

    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 📥 Step 1: Upload Scraped Games")
        st.info("Run the Python scraper locally on your computer, then upload the generated games_schedule.csv file here.")
        games_file = st.file_uploader("Upload games_schedule.csv", type=["csv"])
        if games_file is not None:
            saved, message = save_games_schedule_upload(games_file)
            if saved:
                st.success(message)
            else:
                st.error(message)
        elif os.path.exists(GAMES_SCHEDULE_PATH):
            st.caption(f"Using existing {GAMES_SCHEDULE_PATH} from a previous upload.")

    with col2:
        st.subheader("Step 2: Generate Schedule")
        if st.button("🧠 Generate Excel Matrix", type="primary", use_container_width=True, disabled=is_approved):
            
            # THE NEW SAFETY LOCK!
            if not os.path.exists(GAMES_SCHEDULE_PATH):
                st.error("🚨 Missing game data! Please upload games_schedule.csv in Step 1 first.")
            else:
                with st.spinner("Running Master Algorithm..."):
                    generated_filename = scheduler.generate_matrix(selected_year, selected_month, all_ptos, all_rdos, holiday_workers)
                    
                if generated_filename and os.path.exists(generated_filename):
                    with open(generated_filename, "rb") as f:
                        excel_data = f.read()
                    conn = get_db_connection()
                    c = conn.cursor()
                    c.execute("""
                        INSERT INTO schedule_status (target_year, target_month, is_approved, excel_file)
                        VALUES (%s, %s, FALSE, %s)
                        ON CONFLICT (target_year, target_month)
                        DO UPDATE SET excel_file = EXCLUDED.excel_file, is_approved = FALSE;
                    """, (selected_year, selected_month, psycopg2.Binary(excel_data)))
                    conn.commit()
                    conn.close()
                    
                    st.success("Draft Matrix generated and safely backed up to Cloud!")
                    st.rerun()
                else:
                    st.error("Failed to generate matrix. Please check inputs.")

    st.markdown("---")
    if db_file_bytes:
        st.markdown(f"### 📊 {calendar.month_name[selected_month]} Schedule Preview")
        try:
            preview_df = pd.read_excel(io.BytesIO(db_file_bytes), index_col=0).fillna("") 
            st.dataframe(preview_df, use_container_width=True)
        except Exception as e:
            st.error("Could not load preview.")
            
        col_down, col_approve = st.columns(2)
        with col_down:
            st.download_button(
                label=f"📥 Download {'Approved' if is_approved else 'Draft'} Matrix", 
                data=db_file_bytes, 
                file_name=f"{'APPROVED_' if is_approved else ''}{calendar.month_name[selected_month]}_{selected_year}_Matrix.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_approve:
            if not is_approved:
                if st.button("✅ Approve & Lock Schedule", type="primary", use_container_width=True):
                    conn = get_db_connection()
                    c = conn.cursor()
                    c.execute("UPDATE schedule_status SET is_approved=TRUE WHERE target_year=%s AND target_month=%s", (selected_year, selected_month))
                    conn.commit()
                    conn.close()
                    st.success("Schedule Approved and Locked!")
                    st.rerun()

        # ==========================================
        # NEW FEATURE: MANUAL EDITS UPLOAD
        # ==========================================
        if not is_approved:
            st.markdown("---")
    if db_file_bytes:
        st.markdown(f"### 📊 {calendar.month_name[selected_month]} Schedule Editor")
        try:
            import io
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Read the current file from the Cloud
            preview_df = pd.read_excel(io.BytesIO(db_file_bytes), index_col=0).fillna("") 
            
            if not is_approved:
                st.info("💡 **LIVE EDITOR:** Double-click any cell below to change a shift. Click 'Save Live Edits' when you are done.")
                
                # THIS IS THE MAGIC STREAMLIT EDITOR!
                edited_df = st.data_editor(preview_df, use_container_width=True)
                
                col_save, col_approve, col_down = st.columns(3)
                
                with col_save:
                    if st.button("💾 Save Live Edits", type="primary", use_container_width=True):
                        # 1. Convert edited data back to Excel in memory
                        output = io.BytesIO()
                        edited_df.to_excel(output, index=True)
                        output.seek(0)
                        
                        # 2. Re-paint the Excel colors so it looks pretty!
                        wb = openpyxl.load_workbook(output)
                        ws = wb.active
                        
                        MagentaFill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')
                        BlackFill = PatternFill(start_color='000000', end_color='000000', fill_type='solid') 
                        OrangeFill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid') 
                        CyanFill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid') 

                        ws.row_dimensions[1].height = 25 
                        for cell in ws[1]:
                            cell.fill = OrangeFill
                            cell.font = Font(bold=True, color='FFFFFF')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
                            
                        ws.column_dimensions['A'].width = 25 
                        for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
                            ws.column_dimensions[col[0].column_letter].width = 10 
                            
                        for row_num in range(2, ws.max_row + 1):
                            for col_num in range(2, ws.max_column + 1):
                                cell = ws.cell(row=row_num, column=col_num)
                                cell_val = str(cell.value).strip() if cell.value else ""
                                
                                if cell_val == "" or cell_val == 'OFF':
                                    cell.fill = BlackFill
                                    cell.font = Font(color='FFFFFF')
                                    cell.value = "" 
                                elif cell_val == 'PTO' or cell_val == 'HOLIDAY':
                                    cell.fill = CyanFill 
                                    cell.font = Font(bold=True, color='000000')
                                else:
                                    cell.fill = MagentaFill # Default color for working shifts
                                    
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
                        
                        # 3. Save the repainted file back to Supabase
                        final_output = io.BytesIO()
                        wb.save(final_output)
                        final_bytes = final_output.getvalue()
                        
                        conn = get_db_connection()
                        c = conn.cursor()
                        c.execute("UPDATE schedule_status SET excel_file=%s WHERE target_year=%s AND target_month=%s", (psycopg2.Binary(final_bytes), selected_year, selected_month))
                        conn.commit()
                        conn.close()
                        st.success("Edits saved successfully!")
                        st.rerun()

                with col_approve:
                    if st.button("✅ Approve & Lock Schedule", use_container_width=True):
                        conn = get_db_connection()
                        c = conn.cursor()
                        c.execute("UPDATE schedule_status SET is_approved=TRUE WHERE target_year=%s AND target_month=%s", (selected_year, selected_month))
                        conn.commit()
                        conn.close()
                        st.success("Schedule Approved and Locked!")
                        st.rerun()
                        
                with col_down:
                    st.download_button(
                        label="📥 Download Draft", 
                        data=db_file_bytes, 
                        file_name=f"DRAFT_{calendar.month_name[selected_month]}_{selected_year}_Matrix.xlsx", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            # If the schedule IS approved, just show the locked table (no editing)
            else:
                st.dataframe(preview_df, use_container_width=True)
                st.download_button(
                    label="📥 Download Approved Matrix", 
                    data=db_file_bytes, 
                    file_name=f"APPROVED_{calendar.month_name[selected_month]}_{selected_year}_Matrix.xlsx", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                
        except Exception as e:
            st.error("Could not load preview.")

        # ==========================================
        # SMART GAME AUTO-ASSIGNER
        # ==========================================
        st.markdown("---")
        st.markdown("### 🎮 Smart Game Auto-Assigner")
        st.info("This tool looks at your generated Schedule Matrix and the scraped games, then automatically assigns the best person based on their shift!")

        if not os.path.exists(GAMES_SCHEDULE_PATH):
            st.warning("⚠️ No games data found! Please upload the games_schedule.csv file in Step 1.")
        else:
            # We use session_state so the table doesn't disappear when you edit a cell
            if "assignments_df" not in st.session_state:
                if st.button("🤖 Auto-Assign Games", type="primary", use_container_width=True):
                    with st.spinner("Calculating optimal assignments..."):
                        games_df = pd.read_csv(GAMES_SCHEDULE_PATH)
                        matrix_df = pd.read_excel(io.BytesIO(db_file_bytes), index_col=0).fillna("")
                        
                        assignment_rows = []
                        daily_workload = {}
                        
                        for _, game in games_df.iterrows():
                            g_date_str = game['Date'] 
                            g_date_obj = datetime.strptime(g_date_str, "%Y-%m-%d")
                            
                            # --- 1. NEW: ONLY SHOW GAMES FOR THE CURRENTLY SELECTED MONTH ---
                            if g_date_obj.month != selected_month or g_date_obj.year != selected_year:
                                continue
                            # ----------------------------------------------------------------
                            
                            g_start_str = str(game['Coverage_Start']).strip()
                            g_end_str = str(game['Coverage_End']).strip()
                            
                            matrix_col = f"{g_date_obj.strftime('%a')} {g_date_obj.day}"
                            
                            g_start_dt = datetime.strptime(f"{g_date_str} {g_start_str}", "%Y-%m-%d %H:%M")
                            
                            # --- INCORPORATED NCAA FB TIME LOGIC ---
                            if not g_end_str or g_end_str.lower() == 'nan':
                                if game.get('Sport') == "CFL":
                                    hours_to_add = 3
                                elif game.get('Sport') == "NCAA FB":
                                    hours_to_add = 3
                                else:
                                    hours_to_add = 2
                                g_end_dt = g_start_dt + timedelta(hours=hours_to_add)
                            else:
                                g_end_dt = datetime.strptime(f"{g_date_str} {g_end_str}", "%Y-%m-%d %H:%M")
                            # ---------------------------------------
                            
                            if g_end_dt < g_start_dt:
                                g_end_dt += timedelta(days=1)
                                
                            # --- 2. NEW: CALCULATE UTC TIMES (+5 HOURS) ---
                            utc_start_dt = g_start_dt + timedelta(hours=5)
                            utc_end_dt = g_end_dt + timedelta(hours=5)
                            # ----------------------------------------------
                            
                            assigned_person = "UNASSIGNED ⚠️"
                            
                            if matrix_col in matrix_df.columns:
                                perfect_matches = []
                                partial_matches = []
                                for staff_name, shift_str in matrix_df[matrix_col].items():
                                    shift_str = str(shift_str).strip()
                                    if shift_str not in ["OFF", "PTO", "HOLIDAY", ""]:
                                        try:
                                            s_start, s_end = shift_str.split(" - ")
                                            shift_start_dt = datetime.strptime(f"{g_date_str} {s_start.strip()}", "%Y-%m-%d %H:%M")
                                            shift_end_dt = datetime.strptime(f"{g_date_str} {s_end.strip()}", "%Y-%m-%d %H:%M")
                                            if shift_end_dt < shift_start_dt:
                                                shift_end_dt += timedelta(days=1)
                                                
                                            # PERFECT MATCH: Shift covers the entire game
                                            if shift_start_dt <= g_start_dt and shift_end_dt >= g_end_dt:
                                                perfect_matches.append(staff_name)
                                            # PARTIAL MATCH: Shift covers the start of the game (for late games)
                                            elif shift_start_dt <= g_start_dt < shift_end_dt:
                                                partial_matches.append(staff_name)
                                        except:
                                            pass
                                    
                                # Prioritize perfect matches, but use partial matches as a safety net!
                                eligible_staff = perfect_matches if perfect_matches else partial_matches
                                
                                if eligible_staff:
                                    if g_date_str not in daily_workload:
                                        daily_workload[g_date_str] = {s: 0 for s in matrix_df.index}
                                    # Balance the workload so one person doesn't get all the late games
                                    best_staff = min(eligible_staff, key=lambda s: daily_workload[g_date_str].get(s, 0))
                                    assigned_person = best_staff
                                    daily_workload[g_date_str][best_staff] += 1
                                    
                            venue = game.get('Venue', "")
                            
                            # --- 3. NEW: OUTPUTTING THE UTC COLUMNS ---
                            assignment_rows.append({
                                "Coverage Start (UTC-5)": g_start_dt.strftime("%m/%d/%Y %H:%M"),
                                "Coverage End (UTC-5)": g_end_dt.strftime("%m/%d/%Y %H:%M"),
                                "Coverage Start (UTC)": utc_start_dt.strftime("%m/%d/%Y %H:%M"),
                                "Coverage End (UTC)": utc_end_dt.strftime("%m/%d/%Y %H:%M"),
                                "Matchup": game['Matchup'],
                                "Event ID/League": game['Sport'],
                                "Venue": venue,
                                "Assigned Name": assigned_person,
                                "QA 1": "",
                                "QA 2": "",
                                "ID": ""
                            })
                            
                        st.session_state.assignments_df = pd.DataFrame(assignment_rows)
                        st.rerun()
            
            if "assignments_df" in st.session_state:
                st.success("Assignments generated! You can double-click any cell to manually override the AI before downloading.")
                
                # The interactive editor!
                edited_assignments = st.data_editor(st.session_state.assignments_df, use_container_width=True)
                
                # Create the downloadable Excel file
                output = io.BytesIO()
                edited_assignments.to_excel(output, index=False)
                output.seek(0)
                
                col_reset, col_down = st.columns(2)
                with col_down:
                    st.download_button(
                        label="📥 Download Game Assignments (.xlsx)",
                        data=output,
                        file_name=f"Game_Assignments_{calendar.month_name[selected_month]}_{selected_year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                with col_reset:
                    if st.button("🔄 Recalculate Assignments", use_container_width=True):
                        del st.session_state.assignments_df
                        st.rerun()



