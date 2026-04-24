import calendar
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import psycopg2
import streamlit as st

TEAM_MEMBERS = ["Alberto Salazar", "Camilo Buritica", "Emilio Gonzalez", "Juan Camilo Correa", "Simon Mejia", "Brayan Carlosama"]

def generate_matrix(YEAR, MONTH, PTO_REQUESTS, REQUESTED_DAYS_OFF, HOLIDAYS):
    MAX_WEEKLY_HOURS = 44
    MIN_REST_HOURS = 12
    MAX_CONSECUTIVE_DAYS = 6
    FIXED_DAYS_OFF = {"Alberto Salazar": 1} 

    def get_db_connection():
        return psycopg2.connect(st.secrets["connections"]["supabase"]["url"])

    def load_carry_over_data(target_year, target_month):
        if target_month == 1:
            prev_m, prev_y = 12, target_year - 1
        else:
            prev_m, prev_y = target_month - 1, target_year
            
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""
            SELECT staff_member, weekly_hours, days_worked_this_week, has_had_short_day, 
                   last_shift_end_time, consecutive_days, last_worked_date, night_shifts_this_month 
            FROM carry_over_stats 
            WHERE target_year=%s AND target_month=%s
        """, (prev_y, prev_m))
        rows = c.fetchall()
        conn.close()
        
        if rows:
            data = {}
            for r in rows:
                name, w_h, d_w, h_h_s, l_s_e, c_d, l_w_d, n_s = r
                data[name] = {
                    "weekly_hours": w_h,
                    "days_worked_this_week": d_w,
                    "has_had_short_day": h_h_s,
                    "last_shift_end_time": datetime.strptime(l_s_e, "%Y-%m-%d %H:%M:%S") if l_s_e else None,
                    "consecutive_days": c_d,
                    "last_worked_date": datetime.strptime(l_w_d, "%Y-%m-%d") if l_w_d else None,
                    "night_shifts_this_month": n_s if n_s else 0
                }
            return data
                
        elif target_year == 2026 and target_month == 5:
            return {
                "Alberto Salazar": {"weekly_hours": 27, "days_worked_this_week": 3, "has_had_short_day": False, "last_shift_end_time": datetime(2026, 4, 30, 18, 0), "consecutive_days": 2, "last_worked_date": datetime(2026, 4, 30), "night_shifts_this_month": 0},
                "Camilo Buritica": {"weekly_hours": 35, "days_worked_this_week": 4, "has_had_short_day": True, "last_shift_end_time": datetime(2026, 4, 30, 20, 0), "consecutive_days": 4, "last_worked_date": datetime(2026, 4, 30), "night_shifts_this_month": 0},
                "Emilio Gonzalez": {"weekly_hours": 26, "days_worked_this_week": 3, "has_had_short_day": True, "last_shift_end_time": datetime(2026, 4, 30, 18, 0), "consecutive_days": 1, "last_worked_date": datetime(2026, 4, 30), "night_shifts_this_month": 0},
                "Juan Camilo Correa": {"weekly_hours": 35, "days_worked_this_week": 4, "has_had_short_day": True, "last_shift_end_time": datetime(2026, 4, 30, 18, 0), "consecutive_days": 4, "last_worked_date": datetime(2026, 4, 30), "night_shifts_this_month": 0},
                "Simon Mejia": {"weekly_hours": 35, "days_worked_this_week": 4, "has_had_short_day": True, "last_shift_end_time": datetime(2026, 4, 30, 16, 0), "consecutive_days": 4, "last_worked_date": datetime(2026, 4, 30), "night_shifts_this_month": 0},
                "Brayan Carlosama": {"weekly_hours": 36, "days_worked_this_week": 4, "has_had_short_day": False, "last_shift_end_time": datetime(2026, 4, 30, 18, 0), "consecutive_days": 4, "last_worked_date": datetime(2026, 4, 30), "night_shifts_this_month": 0}
            }
        else:
            return {name: {"weekly_hours": 0, "days_worked_this_week": 0, "has_had_short_day": False, "last_shift_end_time": None, "consecutive_days": 0, "last_worked_date": None, "night_shifts_this_month": 0} for name in TEAM_MEMBERS}

    team_stats = load_carry_over_data(YEAR, MONTH)

    for name in TEAM_MEMBERS:
        if "night_shifts_this_month" not in team_stats[name]:
            team_stats[name]["night_shifts_this_month"] = 0

    def load_schedule_data(filename="games_schedule.csv"):
        try: return pd.read_csv(filename)
        except FileNotFoundError: return pd.DataFrame(columns=['Date', 'Sport', 'Matchup', 'Coverage_Start', 'Coverage_End'])

    games_data = load_schedule_data()

    def get_actual_shift_times(cov_start, cov_end, is_short_day, current_dt):
        shift_hours = 8 if is_short_day else 9
        shift_end = cov_end
        if shift_end.minute > 0:
            shift_end = shift_end.replace(minute=0) + timedelta(hours=1)
        shift_start = shift_end - timedelta(hours=shift_hours)
        if shift_start > cov_start:
            shift_start = cov_start.replace(minute=0)
            shift_end = shift_start + timedelta(hours=shift_hours)
        min_start_hour = 9 if current_dt.weekday() >= 5 else 8
        min_start = current_dt.replace(hour=min_start_hour, minute=0, second=0, microsecond=0)
        if shift_start < min_start:
            shift_start = min_start
            shift_end = shift_start + timedelta(hours=shift_hours)
        return shift_start, shift_end

    def needs_short_day(name, current_dt, short_days_assigned_today):
        if current_dt.weekday() >= 5: return False 
        stats = team_stats[name]
        if stats["has_had_short_day"]: return False
        if current_dt.weekday() == 4 or stats["days_worked_this_week"] >= 4: return True
        quota = 2 if current_dt.weekday() == 2 else 1
        return short_days_assigned_today < quota

    def needs_weekend_catchup(name, current_dt):
        stats = team_stats[name]
        if current_dt.weekday() == 5: 
            if stats["days_worked_this_week"] <= 3: return True
            if stats["days_worked_this_week"] == 4:
                tomorrow_str = (current_dt + timedelta(days=1)).strftime("%Y-%m-%d")
                cant_work_sun = (name in REQUESTED_DAYS_OFF and tomorrow_str in REQUESTED_DAYS_OFF[name]) or \
                                (name in PTO_REQUESTS and tomorrow_str in PTO_REQUESTS[name]) or \
                                (name in FIXED_DAYS_OFF and FIXED_DAYS_OFF[name] == 6)
                return cant_work_sun
            return False
        elif current_dt.weekday() == 6: 
            return stats["days_worked_this_week"] < 5
        return False

    def is_eligible_for_shift(name, current_dt, shift_start_dt, shift_hours):
        stats = team_stats[name]
        date_str = current_dt.strftime("%Y-%m-%d")
        
        if name in REQUESTED_DAYS_OFF and date_str in REQUESTED_DAYS_OFF[name]: return False
        if name in FIXED_DAYS_OFF and FIXED_DAYS_OFF[name] == current_dt.weekday(): return False
        if stats["days_worked_this_week"] >= 5: return False
        if stats["weekly_hours"] + shift_hours > MAX_WEEKLY_HOURS: return False
        
        effective_consecutive = 0 if stats["last_worked_date"] != current_dt - timedelta(days=1) else stats["consecutive_days"]
        if effective_consecutive >= MAX_CONSECUTIVE_DAYS: return False
            
        if stats["last_shift_end_time"]:
            rest_time = shift_start_dt - stats["last_shift_end_time"]
            if (rest_time.total_seconds() / 3600) < MIN_REST_HOURS: 
                return False
                
        return True

    def evaluate_weekend_demand(current_date):
        monday = current_date - timedelta(days=current_date.weekday())
        sat_date = monday + timedelta(days=5)
        sun_date = monday + timedelta(days=6)
        total_comp_needed = 0
        
        for d_date in [sat_date, sun_date]:
            d_str = d_date.strftime("%Y-%m-%d")
            day_games = games_data[games_data['Date'] == d_str].sort_values(by='Coverage_Start')
            planned = []
            
            for _, g in day_games.iterrows():
                sport = str(g['Sport']).strip().upper()
                pre_game = 1 if 'WNBA' in sport else 2
                dur = 2 if 'WNBA' in sport else 3
                
                g_actual = datetime.strptime(f"{d_str} {g['Coverage_Start']}", "%Y-%m-%d %H:%M")
                g_cov_start = g_actual - timedelta(hours=pre_game)
                g_cov_end = g_actual + timedelta(hours=dur)
                
                placed = False
                for s in planned:
                    if s['count'] < 4: 
                        n_s = min(s['start'], g_cov_start)
                        n_e = max(s['end'], g_cov_end)
                        if (n_e - n_s).total_seconds() <= 9 * 3600:
                            s['start'] = n_s
                            s['end'] = n_e
                            s['count'] += 1
                            placed = True
                            break
                if not placed:
                    planned.append({'start': g_cov_start, 'end': g_cov_end, 'count': 1})
                    
            if d_date == sat_date and len(planned) == 0:
                total_comp_needed += 1
            else:
                total_comp_needed += len(planned)
                
        rdo_this_week = 0
        for name in TEAM_MEMBERS:
            for i in range(5):
                chk_date = (monday + timedelta(days=i)).strftime("%Y-%m-%d")
                if name in REQUESTED_DAYS_OFF and chk_date in REQUESTED_DAYS_OFF[name]:
                    rdo_this_week += 1
                    break
                    
        return max(0, total_comp_needed - len(FIXED_DAYS_OFF) - rdo_this_week)

    master_schedule_data = []
    num_days = calendar.monthrange(YEAR, MONTH)[1] 
    early_shifts_count = {name: 0 for name in TEAM_MEMBERS}

    comp_queue = [m for m in TEAM_MEMBERS if m not in FIXED_DAYS_OFF]
    comp_idx = 0
    comp_days_to_give = 0
    comp_days_given = 0

    def reset_weekly_stats():
        for name in TEAM_MEMBERS:
            team_stats[name]["weekly_hours"] = 0
            team_stats[name]["days_worked_this_week"] = 0
            team_stats[name]["has_had_short_day"] = False

    for day in range(1, num_days + 1):
        current_date = datetime(YEAR, MONTH, day)
        date_str = current_date.strftime("%Y-%m-%d")
        display_date = f"{current_date.strftime('%a')} {day}"
        is_weekday = current_date.weekday() < 5 
        
        if current_date.weekday() == 0 or day == 1:
            if current_date.weekday() == 0:
                reset_weekly_stats()
                comp_days_given = 0
            comp_days_to_give = evaluate_weekend_demand(current_date)
            
        working_today = set()
        short_days_assigned_today = 0 

        if date_str in HOLIDAYS:
            workers_on_holiday = HOLIDAYS[date_str]
            for name in TEAM_MEMBERS:
                if name not in workers_on_holiday:
                    working_today.add(name)
                    team_stats[name]["days_worked_this_week"] += 1
                    if not team_stats[name]["has_had_short_day"] and is_weekday:
                        pto_hours = 8
                        team_stats[name]["has_had_short_day"] = True
                    else:
                        pto_hours = 9
                    team_stats[name]["weekly_hours"] += pto_hours
                    team_stats[name]["consecutive_days"] = 0 
                    master_schedule_data.append({
                        "Date_Display": display_date, "Staff Member": name,
                        "Shift": "HOLIDAY", "Color_Key": "PTO"
                    })

        for name in TEAM_MEMBERS:
            if name in PTO_REQUESTS and date_str in PTO_REQUESTS[name]:
                if name not in working_today: 
                    working_today.add(name)
                    team_stats[name]["days_worked_this_week"] += 1
                    if not team_stats[name]["has_had_short_day"] and is_weekday:
                        pto_hours = 8
                        team_stats[name]["has_had_short_day"] = True
                    else:
                        pto_hours = 9
                    team_stats[name]["weekly_hours"] += pto_hours
                    team_stats[name]["consecutive_days"] = 0 
                    master_schedule_data.append({
                        "Date_Display": display_date, "Staff Member": name,
                        "Shift": "PTO", "Color_Key": "PTO"
                    })
        
        if current_date.weekday() in [2, 3, 4] and comp_days_given < comp_days_to_give:
            days_left_in_comp_window = 5 - current_date.weekday() 
            comp_days_needed_today = (comp_days_to_give - comp_days_given) // days_left_in_comp_window
            if (comp_days_to_give - comp_days_given) % days_left_in_comp_window > 0:
                comp_days_needed_today += 1
                
            for _ in range(comp_days_needed_today):
                attempts = 0
                while attempts < len(comp_queue):
                    name_to_force_off = comp_queue[comp_idx]
                    comp_idx = (comp_idx + 1) % len(comp_queue)
                    attempts += 1
                    
                    is_working_holiday = (date_str in HOLIDAYS and name_to_force_off in HOLIDAYS[date_str])
                    
                    if name_to_force_off not in working_today and (name_to_force_off not in FIXED_DAYS_OFF or FIXED_DAYS_OFF[name_to_force_off] != current_date.weekday()) and not is_working_holiday:
                        working_today.add(name_to_force_off) 
                        comp_days_given += 1
                        break

        late_coverage_met = False
        todays_games_df = games_data[games_data['Date'] == date_str].sort_values(by='Coverage_Start')
        planned_shifts = []
        
        if is_weekday:
            late_games = []
            for _, g in todays_games_df.iterrows():
                sport = str(g['Sport']).strip().upper()
                pre_game = 1 if 'WNBA' in sport else 2
                dur = 2 if 'WNBA' in sport else 3
                
                g_actual = datetime.strptime(f"{date_str} {g['Coverage_Start']}", "%Y-%m-%d %H:%M")
                g_cov_start = g_actual - timedelta(hours=pre_game)
                g_cov_end = g_actual + timedelta(hours=dur)
                
                b_end = current_date.replace(hour=18, minute=0, second=0, microsecond=0)
                if g_cov_end > b_end:
                    late_games.append((g_cov_start, g_cov_end))
                    
            if late_games:
                latest_cov_end = max([g[1] for g in late_games])
                earliest_cov_start = min([g[0] for g in late_games])
                planned_shifts.append({'start': earliest_cov_start, 'end': latest_cov_end, 'count': 1})
                
        else:
            for _, g in todays_games_df.iterrows():
                sport = str(g['Sport']).strip().upper()
                pre_game = 1 if 'WNBA' in sport else 2
                dur = 2 if 'WNBA' in sport else 3
                
                g_actual = datetime.strptime(f"{date_str} {g['Coverage_Start']}", "%Y-%m-%d %H:%M")
                g_cov_start = g_actual - timedelta(hours=pre_game)
                g_cov_end = g_actual + timedelta(hours=dur)
                
                placed = False
                for s in planned_shifts:
                    if s['count'] < 4: 
                        n_s = min(s['start'], g_cov_start)
                        n_e = max(s['end'], g_cov_end)
                        if (n_e - n_s).total_seconds() <= 9 * 3600:
                            s['start'] = n_s
                            s['end'] = n_e
                            s['count'] += 1
                            placed = True
                            break
                if not placed:
                    planned_shifts.append({'start': g_cov_start, 'end': g_cov_end, 'count': 1})
                
        for s in planned_shifts:
            assignment_order = sorted(TEAM_MEMBERS, key=lambda n: (
                0 if (current_date.weekday() in [5, 6] and needs_weekend_catchup(n, current_date)) else 1,
                team_stats[n]["night_shifts_this_month"],
                team_stats[n]["weekly_hours"]
            ))
            
            for name in assignment_order:
                if name in working_today: continue
                
                is_short = needs_short_day(name, current_date, short_days_assigned_today)
                shift_start_dt, shift_end_dt = get_actual_shift_times(s['start'], s['end'], is_short, current_date)
                shift_hours = 8 if is_short else 9
                
                if is_eligible_for_shift(name, current_date, shift_start_dt, shift_hours):
                    working_today.add(name)
                    team_stats[name]["weekly_hours"] += shift_hours
                    team_stats[name]["days_worked_this_week"] += 1
                    if is_short: 
                        team_stats[name]["has_had_short_day"] = True
                        short_days_assigned_today += 1
                    team_stats[name]["last_shift_end_time"] = shift_end_dt
                    
                    if team_stats[name]["last_worked_date"] == current_date - timedelta(days=1): team_stats[name]["consecutive_days"] += 1
                    else: team_stats[name]["consecutive_days"] = 1
                    team_stats[name]["last_worked_date"] = current_date
                    
                    if shift_end_dt.hour >= 20 or shift_end_dt.hour < 6: 
                        late_coverage_met = True 
                        team_stats[name]["night_shifts_this_month"] += 1
                        
                    master_schedule_data.append({
                        "Date_Display": display_date, "Staff Member": name,
                        "Shift": f"{shift_start_dt.strftime('%H')}:00 - {shift_end_dt.strftime('%H')}:00",
                        "Color_Key": "LATE_GAME" if (shift_end_dt.hour >= 20 or shift_end_dt.hour < 6) else "BASELINE"
                    })
                    break 

        if is_weekday:
            if current_date.weekday() in [3, 4] and not late_coverage_met:
                late_assigned = False
                assignment_order = sorted(TEAM_MEMBERS, key=lambda n: (team_stats[n]["night_shifts_this_month"], team_stats[n]["weekly_hours"]))
                
                for name in assignment_order:
                    if name in working_today: continue 
                    
                    is_short = needs_short_day(name, current_date, short_days_assigned_today)
                    shift_hours = 8 if is_short else 9
                    start_hour = 20 - shift_hours 
                    shift_start_dt = current_date.replace(hour=start_hour, minute=0, second=0, microsecond=0)
                    
                    if is_eligible_for_shift(name, current_date, shift_start_dt, shift_hours):
                        shift_end_dt = shift_start_dt + timedelta(hours=shift_hours)
                        master_schedule_data.append({"Date_Display": display_date, "Staff Member": name, "Shift": f"{shift_start_dt.strftime('%H')}:00 - {shift_end_dt.strftime('%H')}:00", "Color_Key": "LATE_GAME"})
                        team_stats[name]["weekly_hours"] += shift_hours
                        team_stats[name]["days_worked_this_week"] += 1
                        team_stats[name]["night_shifts_this_month"] += 1
                        if is_short: 
                            team_stats[name]["has_had_short_day"] = True
                            short_days_assigned_today += 1
                        team_stats[name]["last_shift_end_time"] = shift_end_dt
                        if team_stats[name]["last_worked_date"] == current_date - timedelta(days=1): team_stats[name]["consecutive_days"] += 1
                        else: team_stats[name]["consecutive_days"] = 1
                        team_stats[name]["last_worked_date"] = current_date
                        working_today.add(name)
                        late_assigned = True
                        break

            early_assigned = False
            assignment_order = sorted(TEAM_MEMBERS, key=lambda n: (early_shifts_count[n], team_stats[n]["weekly_hours"]))
            for name in assignment_order:
                if name not in working_today:
                    is_short = needs_short_day(name, current_date, short_days_assigned_today)
                    shift_hours = 8 if is_short else 9
                    shift_start_dt = current_date.replace(hour=8, minute=0, second=0, microsecond=0)
                    
                    if is_eligible_for_shift(name, current_date, shift_start_dt, shift_hours):
                        shift_end_dt = shift_start_dt + timedelta(hours=shift_hours)
                        master_schedule_data.append({"Date_Display": display_date, "Staff Member": name, "Shift": f"{shift_start_dt.strftime('%H')}:00 - {shift_end_dt.strftime('%H')}:00", "Color_Key": "BASELINE"})
                        team_stats[name]["weekly_hours"] += shift_hours
                        team_stats[name]["days_worked_this_week"] += 1
                        if is_short: 
                            team_stats[name]["has_had_short_day"] = True
                            short_days_assigned_today += 1
                        team_stats[name]["last_shift_end_time"] = shift_end_dt
                        if team_stats[name]["last_worked_date"] == current_date - timedelta(days=1): team_stats[name]["consecutive_days"] += 1
                        else: team_stats[name]["consecutive_days"] = 1
                        team_stats[name]["last_worked_date"] = current_date
                        working_today.add(name)
                        
                        early_shifts_count[name] += 1 # <-- WE RECORD THAT THEY TOOK THE SHIFT HERE
                        early_assigned = True
                        break

            assignment_order = sorted(TEAM_MEMBERS, key=lambda n: team_stats[n]["weekly_hours"])
            for name in assignment_order:
                if name not in working_today:
                    is_short = needs_short_day(name, current_date, short_days_assigned_today)
                    shift_hours = 8 if is_short else 9
                    
                    for base_start in range(9, 13):
                        shift_start_hour = base_start
                        if is_short and (base_start + 9) >= 20: shift_start_hour += 1 
                            
                        shift_start_dt = current_date.replace(hour=shift_start_hour, minute=0, second=0, microsecond=0)
                        shift_end_dt = shift_start_dt + timedelta(hours=shift_hours)
                        
                        if is_eligible_for_shift(name, current_date, shift_start_dt, shift_hours):
                            master_schedule_data.append({"Date_Display": display_date, "Staff Member": name, "Shift": f"{shift_start_dt.strftime('%H')}:00 - {shift_end_dt.strftime('%H')}:00", "Color_Key": "BASELINE"})
                            team_stats[name]["weekly_hours"] += shift_hours
                            team_stats[name]["days_worked_this_week"] += 1
                            if is_short: 
                                team_stats[name]["has_had_short_day"] = True
                                short_days_assigned_today += 1
                            team_stats[name]["last_shift_end_time"] = shift_end_dt
                            if team_stats[name]["last_worked_date"] == current_date - timedelta(days=1): team_stats[name]["consecutive_days"] += 1
                            else: team_stats[name]["consecutive_days"] = 1
                            team_stats[name]["last_worked_date"] = current_date
                            working_today.add(name)
                            break 

        if current_date.weekday() in [5, 6]:
            assignment_order = sorted(TEAM_MEMBERS, key=lambda n: team_stats[n]["weekly_hours"])
            for name in assignment_order:
                if name not in working_today and needs_weekend_catchup(name, current_date):
                    shift_hours = 9
                    for base_start in range(9, 13):
                        shift_start_dt = current_date.replace(hour=base_start, minute=0, second=0, microsecond=0)
                        shift_end_dt = shift_start_dt + timedelta(hours=shift_hours)
                        
                        if is_eligible_for_shift(name, current_date, shift_start_dt, shift_hours):
                            master_schedule_data.append({"Date_Display": display_date, "Staff Member": name, "Shift": f"{shift_start_dt.strftime('%H')}:00 - {shift_end_dt.strftime('%H')}:00", "Color_Key": "BASELINE"})
                            team_stats[name]["weekly_hours"] += shift_hours
                            team_stats[name]["days_worked_this_week"] += 1
                            team_stats[name]["last_shift_end_time"] = shift_end_dt
                            if team_stats[name]["last_worked_date"] == current_date - timedelta(days=1): team_stats[name]["consecutive_days"] += 1
                            else: team_stats[name]["consecutive_days"] = 1
                            team_stats[name]["last_worked_date"] = current_date
                            working_today.add(name)
                            break

        if current_date.weekday() == 5 and len(working_today) == 0:
            assignment_order = sorted(TEAM_MEMBERS, key=lambda n: team_stats[n]["weekly_hours"])
            for name in assignment_order:
                if is_eligible_for_shift(name, current_date, current_date.replace(hour=9, minute=0, second=0), 9):
                    master_schedule_data.append({"Date_Display": display_date, "Staff Member": name, "Shift": "09:00 - 18:00", "Color_Key": "BASELINE"})
                    team_stats[name]["weekly_hours"] += 9
                    team_stats[name]["days_worked_this_week"] += 1
                    team_stats[name]["last_shift_end_time"] = current_date.replace(hour=18, minute=0, second=0)
                    if team_stats[name]["last_worked_date"] == current_date - timedelta(days=1): team_stats[name]["consecutive_days"] += 1
                    else: team_stats[name]["consecutive_days"] = 1
                    team_stats[name]["last_worked_date"] = current_date
                    working_today.add(name)
                    break

    conn = get_db_connection()
    c = conn.cursor()
    for name, stats in team_stats.items():
        l_s_e = stats["last_shift_end_time"].strftime("%Y-%m-%d %H:%M:%S") if stats["last_shift_end_time"] else None
        l_w_d = stats["last_worked_date"].strftime("%Y-%m-%d") if stats["last_worked_date"] else None
        
        c.execute("""
            INSERT INTO carry_over_stats (target_year, target_month, staff_member, weekly_hours, days_worked_this_week, has_had_short_day, last_shift_end_time, consecutive_days, last_worked_date, night_shifts_this_month)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (target_year, target_month, staff_member)
            DO UPDATE SET
                weekly_hours = EXCLUDED.weekly_hours,
                days_worked_this_week = EXCLUDED.days_worked_this_week,
                has_had_short_day = EXCLUDED.has_had_short_day,
                last_shift_end_time = EXCLUDED.last_shift_end_time,
                consecutive_days = EXCLUDED.consecutive_days,
                last_worked_date = EXCLUDED.last_worked_date,
                night_shifts_this_month = EXCLUDED.night_shifts_this_month;
        """, (YEAR, MONTH, name, stats["weekly_hours"], stats["days_worked_this_week"], stats["has_had_short_day"], l_s_e, stats["consecutive_days"], l_w_d, stats["night_shifts_this_month"]))

    conn.commit()
    conn.close()

    df_raw = pd.DataFrame(master_schedule_data)
    if df_raw.empty:
        return None
        
    df_matrix = df_raw.pivot(index='Staff Member', columns='Date_Display', values='Shift')
    df_matrix = df_matrix.fillna('OFF')

    all_month_dates = [f"{datetime(YEAR, MONTH, d).strftime('%a')} {d}" for d in range(1, num_days + 1)]
    df_matrix = df_matrix.reindex(columns=all_month_dates, fill_value='OFF')

    output_filename = f"{calendar.month_name[MONTH]}_{YEAR}_Matrix_Schedule.xlsx"
    df_matrix.to_excel(output_filename, index=True)

    wb = openpyxl.load_workbook(output_filename)
    ws = wb.active

    MagentaFill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')
    LimeGreenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid') 
    BlackFill = PatternFill(start_color='000000', end_color='000000', fill_type='solid') 
    OrangeFill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid') 
    CyanFill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid') 

    ws.row_dimensions[1].height = 25 
    for cell in ws[1]:
        cell.fill = OrangeFill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))
        
    ws.column_dimensions['A'].width = 25 
    for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
        ws.column_dimensions[col[0].column_letter].width = 10 
        
    for row_num in range(2, ws.max_row + 1):
        staff_name = ws[f'A{row_num}'].value 
        for col_num in range(2, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            date_display = ws.cell(row=1, column=col_num).value 
            
            if cell.value == 'OFF':
                cell.fill = BlackFill
                cell.font = Font(color='FFFFFF')
                cell.value = "" 
            elif cell.value == 'PTO':
                cell.fill = CyanFill 
                cell.font = Font(bold=True, color='000000')
            elif cell.value == 'HOLIDAY':
                cell.fill = CyanFill 
                cell.font = Font(bold=True, color='000000')
            else:
                try:
                    row_lookup = df_raw[(df_raw['Staff Member'] == staff_name) & (df_raw['Date_Display'] == date_display)].iloc[0]
                    cell.fill = LimeGreenFill if row_lookup['Color_Key'] == 'LATE_GAME' else MagentaFill
                except IndexError: cell.fill = MagentaFill 
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(border_style="thin"), left=Side(border_style="thin"), right=Side(border_style="thin"), bottom=Side(border_style="thin"))

    wb.save(output_filename)
    return output_filename
